from django.shortcuts import render, redirect
from login.models import Student,StudentProfile,Teacher,TeacherProfile, User
from .models import Semester, Subject, Topic, Question, Choice
from .forms import ExcelFileUpload,QuestionForm,ChoiceForm,QuestionType
from django.contrib.auth import decorators
from django.forms import modelform_factory, modelformset_factory
import openpyxl
from django import forms
# from django

def index(request):
    student=StudentProfile.objects.get(username=request.user.username)
    semesters=Semester.objects.filter(semester__lte=student.semester)
    if request.GET.get('semester'):
        url = 'subject/'+str(request.GET.get('semester'))
        return redirect(url)
    return render(request, 'takedata/semester.html', {'semesters': semesters})


def subject(request, semester):
    if request.user.role == 'TEACHER':
        teacher=TeacherProfile.objects.get(username=request.user.username)
        subs = teacher.subjects
        subjects=Subject.objects.none()
        for sub in subs:
            subject=Subject.objects.filter(subject=sub)
            subjects=subjects|subject
        return render(request, 'takedata/subject.html', {'subjects': subjects})
    else:
        sem = Semester.objects.get(semester=semester)
        if request.user.dept == 'DMACS':
            subjects=Subject.objects.filter(subject__icontains='UCSH')
        else:
            subjects=Subject.objects.filter(subject__icontains='UBBA')
        subjects = subjects.filter(semester=sem)
    
        return render(request, 'takedata/subject.html', {'semester': sem, 'subjects': subjects})


def topic(request, semester, subject):
    sem = Semester.objects.get(semester=semester)
    sub = Subject.objects.get(subject=subject)
    topics = Topic.objects.filter(semester=sem, subject=sub)
    return render(request, 'takedata/topic.html', {'semester': sem, 'subject': sub, 'topics': topics})

def question_type(request, semester, subject, topic):
    typeform=QuestionType()
    top=Topic.objects.get(topic=topic)
    user=User.objects.get(username=request.user.username)
    print('out')
    if request.method=='POST':
        print('in')
        type=request.POST.get('question_type')
        print(type)
        question=Question(type=type,topic=top,marks=1,username=user.username)
        question.save()
        if question.type=='MCQs':
            for i in range(4):
                choice=Choice(question=question)
                choice.save()
        elif question.type=='Fill in the Blanks':
            choice = Choice(question=question,is_correct=True)
            choice.save()
        else:
            choice1 = Choice(question=question,choice='True')
            choice2 = Choice(question=question,choice='False')
            choice1.save()
            choice2.save()
        return redirect('question',id=question.id)
    return render(request,'takedata/question_type.html',{'typeform':typeform,'subject':subject,'semester':semester})

def question(request,id):
    question = Question.objects.get(id=id)
    ChoiceFormSet = forms.modelformset_factory(Choice, form=ChoiceForm, extra=0)
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        formset = ChoiceFormSet(request.POST, queryset=question.choice_set.all())
        if request.POST.get('submit')=='Cancel':
                 print('hello')
                 question.delete()
        else:
            if form.is_valid():
                form.save()
                question.save()
            if formset.is_valid():
                formset.save()
        topic=Topic.objects.get(topic=question.topic)
        return redirect('question_type',subject=topic.subject.subject,semester=topic.semester.semester,topic=topic.topic)
    else:
        form = QuestionForm(instance=question)
        formset = ChoiceFormSet(queryset=question.choice_set.all())
    return render(request, 'takedata/question.html', {'form': form, 'formset': formset})


def read_subject_data(subject_file):

    wb = openpyxl.load_workbook(subject_file)
    sheet = wb.active
    headers = []
    data = []
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            headers.append(cell.value)

    for row in sheet.iter_rows(min_row=2):
        row_data = {}
        for i, cell in enumerate(row):
            row_data[headers[i]] = cell.value
        data.append(row_data)

    for item in data:
        sub=item['subject']
        sem = Semester.objects.get(semester=int(sub[-3]))
        subject, created = Subject.objects.get_or_create(
            subject=item['subject'],
            semester=sem,
        )
        if not created:
            continue
    return None



def read_topic_data(topic_file):
    wb = openpyxl.load_workbook(topic_file)
    semester = ['SEMESTER-1', 'SEMESTER-2', 'SEMESTER-3',
                'SEMESTER-4', 'SEMESTER-5', 'SEMESTER-6']
    for sem in semester:
        sheet = wb[sem]
        headers = []
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                headers.append(cell.value)

        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for i, cell in enumerate(row):
                row_data[headers[i]] = cell.value
            data.append(row_data)

        for item in data:
            subject=item['subject']
            s = Semester.objects.get(semester=int(subject[-3]))
            sub =Subject.objects.get(subject=item['subject'])
            try:
                topic = Topic.objects.get(topic=item['topic'], subject=sub)
            except Topic.DoesNotExist:
                Topic.objects.create(
                    semester=s,
                    subject=sub,
                    topic=item['topic'],
        )
    return None




def take_subject(request):
    if request.method == 'POST':
        form = ExcelFileUpload(request.POST, request.FILES)
        if form.is_valid():
            excel=form.cleaned_data['file']
            read_subject_data(excel)
            return redirect('/admin')
    else:
        form = ExcelFileUpload()
    return render(request, 'takedata/take_subjects.html', {'form': form})


def take_topic(request):
    if request.method == 'POST':
        form = ExcelFileUpload(request.POST, request.FILES)
        if form.is_valid():
            excel=form.cleaned_data['file']
            print(excel)
            read_topic_data(excel)
            return redirect('/admin')
    else:
        form = ExcelFileUpload()
    return render(request, 'takedata/take_topic.html', {'form': form})
